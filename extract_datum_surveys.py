import json
import re
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
INPUT_XLSX = BASE_DIR / "Datum_desagregado.xlsx"
OUTPUT_JSON = BASE_DIR / "datos_datum.json"


# Canonical month keys used by the dashboard (DATUM-specific timeline)
SHEET_TO_MONTH = {
    "datum-diciembre": "Diciembre 2025",
    "datum-enero 1": "Enero (1) 2026",
    "datum-enero 2": "Enero (2) 2026",
    "datum-febrero": "Febrero 2026",
}


# DATUM names vary across tabs (party suffixes, punctuation, abbreviations).
# Normalize to the same canonical names used in the UI whenever possible.
NAME_MAP = {
    "rafael lópez a. (rp)": "Rafael López Aliaga",
    "rafael lópez aliaga": "Rafael López Aliaga",
    "keiko fujimori (fp)": "Keiko Fujimori",
    "keiko fujimori": "Keiko Fujimori",
    "mario vizcarra (pp)": "Mario Vizcarra",
    "mario vizcarra": "Mario Vizcarra",
    "carlos álvarez (ppt)": "Carlos Álvarez",
    "carlos álvarez": "Carlos Álvarez",
    "alfonso lópez chau (an)": "Alfonso López Chau",
    "alfonso lópez chau": "Alfonso López Chau",
    "yonhy lescano (cp)": "Yonhy Lescano",
    "yonhy lescano": "Yonhy Lescano",
    "césar acuña (app)": "César Acuña",
    "césar acuña": "César Acuña",
    "ricardo belmont (obras)": "Ricardo Belmont",
    "ricardo belmont": "Ricardo Belmont",
    "josé luna gálvez (pp)": "José Luna Gálvez",
    "josé luna gálvez": "José Luna Gálvez",
    "george forsyth (sp)": "George Forsyth",
    "george forsyth": "George Forsyth",
    "fernando olivera (fe)": "Fernando Olivera",
    "fernando olivera": "Fernando Olivera",
    "robert chiabra (un)": "Roberto Chiabra",
    "roberto chiabra": "Roberto Chiabra",
    "carlos espá (sc)": "Carlos Espá",
    "carlos espá": "Carlos Espá",
    "rafael belaunde (lp)": "Rafael Belaunde",
    "jorge nieto (bg)": "Jorge Nieto",
    "cand. acción popular": "Candidato de Acción Popular",
    "vladimir cerrón (pl)": "Vladimir Cerrón",
    "vladimir cerrón": "Vladimir Cerrón",
    "roberto sánchez": "Roberto Sánchez",
    "enrique valderrama": "Enrique Valderrama",
    "otros candidatos": "Otros",
    "otros": "Otros",
    "ninguno/blanco/viciado": "Blanco/viciado/ninguno",
    "ninguno/ blanco/ viciado": "Blanco/viciado/ninguno",
    "ninguno, blanco, viciado": "Blanco/viciado/ninguno",
    "no sabe": "No precisa",  # DATUM equivalent label
    "no iría a votar": "No iría a votar",
    "no iria a votar": "No iría a votar",
}


SPECIAL_CATEGORIES = {
    "Blanco/viciado/ninguno",
    "Otros",
    "No precisa",
    "No iría a votar",
}


def normalize_sheet_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip()).lower()


def normalize_name(raw_name: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(raw_name).strip())
    return NAME_MAP.get(cleaned.lower(), cleaned)


def to_percent(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value) * 100, 1)
    return None


def build_candidate(row):
    raw_name = row[0]
    if raw_name is None:
        return None

    total = to_percent(row[1])
    if total is None:
        return None

    name = normalize_name(raw_name)

    candidate = {
        "name": name,
        "total": total,
        "gender": {
            "Masculino": to_percent(row[2]),
            "Femenino": to_percent(row[3]),
        },
        "geography": {
            "Lima/Callao": to_percent(row[4]),
            "Norte": to_percent(row[5]),
            "Centro": to_percent(row[6]),
            "Sur": to_percent(row[7]),
            "Oriente": to_percent(row[8]),
            "Urbana": to_percent(row[9]),
            "Rural": to_percent(row[10]),
        },
    }

    # Remove nulls if any malformed row slips through
    candidate["gender"] = {k: v for k, v in candidate["gender"].items() if v is not None}
    candidate["geography"] = {k: v for k, v in candidate["geography"].items() if v is not None}

    return candidate


def sort_candidates(candidates):
    # Keep special categories after candidate names for cleaner legends/data review.
    def key_fn(item):
        is_special = item["name"] in SPECIAL_CATEGORIES
        return (1 if is_special else 0, -item.get("total", 0), item["name"])

    return sorted(candidates, key=key_fn)


def extract():
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    output = {}

    for ws in wb.worksheets:
        sheet_key = normalize_sheet_name(ws.title)
        month_key = SHEET_TO_MONTH.get(sheet_key)
        if not month_key:
            # Skip unexpected tabs instead of failing hard.
            continue

        candidates = []
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, 12)]
            candidate = build_candidate(row)
            if candidate is not None:
                candidates.append(candidate)

        if not candidates:
            continue

        output[month_key] = {
            "metadata": {
                "pollster": "Datum",
                "source_sheet": ws.title.strip(),
            },
            "candidates": sort_candidates(candidates),
        }

    # Force chronological order in final JSON
    ordered = {}
    for month in ["Diciembre 2025", "Enero (1) 2026", "Enero (2) 2026", "Febrero 2026"]:
        if month in output:
            ordered[month] = output[month]

    return ordered


def main():
    data = extract()
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Generado: {OUTPUT_JSON.name} ({len(data)} meses)")
    for month, payload in data.items():
        print(f"- {month}: {len(payload['candidates'])} filas")


if __name__ == "__main__":
    main()
